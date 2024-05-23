import re, requests, base64, os, urllib3
from pathlib import Path
from collections import Counter
from random import randint
from time import sleep
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from loguru import logger
from dotenv import load_dotenv

load_dotenv()
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

DRIVER = webdriver.Chrome()
WAIT = WebDriverWait(DRIVER, timeout=10)

URL_BASE_ACME: str = os.environ['URL_BASE_ACME']
URL_API_MAILTRAP: str = os.environ['URL_API_MAILTRAP']
TOKEN_MAILTRAP: str = os.environ['TOKEN_MAILTRAP']

def create_xlsx_file(wb_name: str, datas: dict) -> None:
    """Gerar arquivo excel a partir de  nome e dados dinâmicos

        Parâmetros:
        wb_name (str):Nome do Workbook
        datas (dict): Dados da planilha

        Retorno:
            None    
    """
    wb = Workbook()
    sheet_names: list = wb.sheetnames
    wb.remove(wb[sheet_names[-1]])
    for key, values in datas.items():
        ws = wb.create_sheet(key)
        list(map(lambda row: ws.append(row), values))
    wb.save(f"{wb_name}.xlsx")

def login(user: str, password: str) -> bool:
    DRIVER.get(f"{URL_BASE_ACME}/login")

    WAIT.until(EC.visibility_of_element_located((By.ID, "email"))).send_keys(user)
    WAIT.until(EC.visibility_of_element_located((By.ID, "password"))).send_keys(password)

    WAIT.until(EC.visibility_of_element_located((By.CLASS_NAME, "btn.btn-primary"))).click()
    WAIT.until(EC.presence_of_element_located((By.ID, "dashmenu")))
    DRIVER.get(f"{URL_BASE_ACME}/work-items")
    logger.info("loged successful")
    return True

def send_email(message: str, file_path: str) -> None:
    with open(file_path, 'rb') as file:
        file_content = file.read()
    encoded_string: str = base64.b64encode(file_content).decode('utf-8')

    headers: dict = {
        'Authorization': f'Bearer {TOKEN_MAILTRAP}', 'Content-Type': 'application/json',
    }

    json_data: dict = {
        'from': {'email': 'mailtrap@demomailtrap.com', 'name': 'Desafio RPA',},
        'to': [{'email': 'wcfontes19@gmail.com',}],
        'subject': 'Resultado Desafio RPA',
        'text': message,
        'category': 'Integration Test',
        "attachments": [
            {
                "content": encoded_string,
                "filename": 'output.xlsx',
                "type": "text",
                "disposition": "attachment"
            }
        ]
    }

    requests.post(f'{URL_API_MAILTRAP}/send', headers=headers, json=json_data, verify=False)
    logger.info("sended email successful")

def extract_itemns() -> dict:
    datas: dict = {'results': []}
    page: int = 1
    while WAIT.until(EC.visibility_of_element_located((By.CLASS_NAME, "table"))):
        DRIVER.get(f"{URL_BASE_ACME}/work-items?page={page}")

        if re.search(r'Oooops, it looks like you have no Work Items!| \
        "User Options" -> "Reset Test Data"', DRIVER.page_source):
            break
        logger.debug(f'get datas of page: {page}')
        
        WAIT.until(EC.visibility_of_element_located((By.CLASS_NAME, "panel.panel-default")))
        rows: list = WAIT.until(EC.presence_of_all_elements_located((By.TAG_NAME, "td")))
        rows: list = list(map(lambda element: element.text, rows))
        list(map(lambda index: datas['results'].append(rows[index+1:index+6]), range(0, len(rows), 6)))
        page += 1
    del page, rows
    DRIVER.quit()
    logger.info("extracted itemns successful")
    return datas

def main():
    login("wcfontes19@gmail.com", "Bot*1234")
    datas: dict = extract_itemns()

    create_xlsx_file('results', datas)
    file_path = str(Path(Path('results.xlsx')).resolve(len(datas['results'])))

    # Updating message with results and send from email
    message: str = "Quantidades\nWI1:\nWI2:\nWI3:\nWI4:\nWI5:\nTotal:"
    for type_code in ['WI1', 'WI2', 'WI3', 'WI4', 'WI5']:
        type_code_qty: int = Counter(list(map(lambda data: 1 if type_code in data else 0 , datas['results'])))[1]
        message: str = re.sub(f'{type_code}:', f'{type_code}: {type_code_qty}\n', message)
    del type_code
    message: str = re.sub(r'Total:', f'Total: {len(datas['results'])}',message)
    send_email(message, file_path)

if __name__ == '__main__':
    main()
